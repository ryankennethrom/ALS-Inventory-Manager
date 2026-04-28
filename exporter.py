from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


class TemplateExporter:
    def __init__(self, template_file, column_mapping):
        """
        template_file: path to .xlsm template
        column_mapping: dict like {"Products": "A12", "Revenue": "B12"}
        """
        self.template_file = template_file
        self.column_mapping = column_mapping

        self.wb = load_workbook(template_file, keep_vba=True)
        self.ws = self.wb.active

    def _parse_cell(self, cell_ref):
        col_letter, row = coordinate_from_string(cell_ref)
        col = column_index_from_string(col_letter)
        return row, col

    def export(self, data_dict, output_file=None):
        """
        Writes values into the template based on column_mapping.

        If value is a list → writes vertically (downward)
        If value is a single value → writes into the mapped cell
        """
        for key, value in data_dict.items():
            if key not in self.column_mapping:
                continue

            start_cell = self.column_mapping[key]
            row, col = self._parse_cell(start_cell)

            if isinstance(value, list):
                # Write vertically (down rows)
                for offset, item in enumerate(value):
                    self.ws.cell(row=row + offset, column=col, value=item)
            else:
                # Single value
                self.ws.cell(row=row, column=col, value=value)

        save_path = output_file if output_file else self.template_file
        self.wb.save(save_path)
