import os
import sys
import subprocess
import sqlite3
import pandas as pd  # for export to Excel
from typing import List, Dict, Any
import DB
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import copy
from pathlib import Path

class RelationInterface:
    def __init__(self, relation_name: str, default_search_text: str, simple_search_field: str,
                 default_filters=dict(), db_path=DB.db_path):
        self.relation_name = relation_name
        self.db_path = db_path
        self.simple_search_field = simple_search_field
        self.filter_dict = default_filters
        self.default_search_text = default_search_text or ""
        self.search_field_text = self.default_search_text
        self.on_search_field_changed(self.search_field_text)
        self.default_filters = copy.deepcopy(self.filter_dict)
        
        self.curr_results = self.on_search_clicked()  # initial load
    
    def set_current_filters_as_default(self):
        self.default_filters = copy.deepcopy(self.filter_dict)
    
    def is_filter_equal(self, other_filter):
        oth_str = str(other_filter)
        cur_str =str(self.filter_dict)
        return oth_str == cur_str

    def is_filter_default(self):
        def_str = str(self.default_filters)
        cur_str =str(self.filter_dict)
        return def_str == cur_str
    
    def on_filter_changed(self, new_filter_dict):
        new_filter_dict = copy.deepcopy(new_filter_dict)
        new_filter_dict["simple_search"] = self.filter_dict["simple_search"]
        self.filter_dict = new_filter_dict

    def on_search_field_changed(self, text):
        self.search_field_text = text
        if text != "":
            self.filter_dict["simple_search"] = {
                        "clauses": [f"{self.simple_search_field} LIKE ?"],
                        "params":[f"{text}%"]
            }
        else:
            self.filter_dict["simple_search"] = {
                        "clauses": [],
                        "params":[]
            }
        
        

    def on_item_clicked(self, item_index: int) -> Dict[str, Any]:
        """Return the item at the given index from the current results"""
        try:
            return self.curr_results[item_index]
        except IndexError:
            raise ValueError(f"Item index {item_index} out of range")
    
    def get_item(self, item_index):
        try:
            return self.curr_results[item_index]
        except IndexError:
            raise ValueError(f"Item index {item_index} out of range")

    def on_item_updated(self, item_index: int, item_details: Dict[str, Any]):
        """Update the row in the database with new details"""
        try:
            item = self.curr_results[item_index]
        except IndexError:
            raise ValueError(f"Item index {item_index} out of range")

        self.validate_date_inputs(item_details)
        # Build UPDATE statement
        set_clause = ", ".join([f"{col}=?" for col in item_details.keys()])
        where_clause = " AND ".join([f"{col}=?" for col in item.keys()])
        params = list(item_details.values()) + list(item.values())
        query = f"UPDATE {self.relation_name} SET {set_clause} WHERE {where_clause}"

        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            cursor.execute(query, params)
            conn.commit()
            if cursor.rowcount == 0:
                raise ValueError(f"Item not found. Someone likely recently updated the item.")

        self.curr_results = self.on_search_clicked()  # refresh

    def on_item_delete_clicked(self, item_index: int):
        """Delete the row at the given index from the database"""
        try:
            item = self.curr_results[item_index]
        except IndexError:
            raise ValueError(f"Item index {item_index} out of range")
        
        where_clause = " AND ".join([f"{col}=?" for col in item.keys()])
        query = f"DELETE FROM {self.relation_name} WHERE {where_clause}"
        params = list(item.values())

        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            cursor.execute(query, params)
            conn.commit()

        self.curr_results = self.on_search_clicked()
    
    def on_create_item_clicked(self, details: dict):
        """Insert a new row into the database. Returns (status, user_message, error_details)."""
        self.validate_date_inputs(details)
        columns = ", ".join(details.keys())
        placeholders = ", ".join(["?"] * len(details))
        params = list(details.values())

        query = f"INSERT INTO {self.relation_name} ({columns}) VALUES ({placeholders})"
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            cursor.execute(query, params)
            conn.commit()

        self.curr_results = self.on_search_clicked()
    
    def validate_date_inputs(self, details):

        def is_valid_date(value: str) -> bool:
            try:
                datetime.strptime(value, "%Y-%m-%d")
                return True
            except (ValueError, TypeError):
                return False

        col_types = DB.get_column_types(self.relation_name)

        for key, value in details.items():
            expected_type = col_types.get(key)

            if expected_type == "date":
                if not is_valid_date(value) and value != "":
                    raise ValueError(
                        f"Invalid date. Make sure {key} has the format YYYY-MM-DD and is a real date."
                    )
        return True

    def get_where_clauses_and_params(self):
        clauses = []
        params = []
        

        for key in self.filter_dict.keys():
            for string_clause in self.filter_dict[key]["clauses"]:
                clauses += [f"{self.relation_name}.{string_clause}"]
            params += self.filter_dict[key]["params"]
        
        def try_int(value):
            try:
                return int(value)
            except (ValueError, TypeError):
                return value

        for i in range(len(params)):
            params[i] = try_int(params[i])

        if clauses:
            return ("WHERE "+ " AND ".join(clauses), params)
        else:
            return ("", [])

    def on_search_clicked(self) -> List[Dict[str, Any]]:
        where_clause, params = self.get_where_clauses_and_params()

        query = f"SELECT * FROM {self.relation_name} "
        
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            cursor.execute(f"{query} {where_clause}", params)
            results = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

        self.curr_results = [dict(zip(columns, row)) for row in results]
        return self.curr_results
    
    def export_as_excel(self, exclude_columns=None, output_path="output.xlsx"):
        if Path(output_path).exists():
            os.remove(output_path)

        if exclude_columns is None:
            exclude_columns = []
       
        query, params = DB.get_expanded_query(self)
        
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON;")
            cursor = conn.cursor()
            cursor.execute(query, params)
            data = cursor.fetchall()

            # Get column names from cursor.description
            columns = [desc[0] for desc in cursor.description]


        df = pd.DataFrame(data, columns=columns)

        df = df.drop(columns=exclude_columns, errors="ignore")

        # ---- Write starting at row 8 (7 empty rows above) ----
        start_row = 7  # zero-indexed for pandas (7 means Excel row 8)
        df.to_excel(output_path, index=False, startrow=start_row)

        wb = load_workbook(output_path)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        if max_row > start_row and max_col > 0:
            header_row = start_row + 1  # Excel row number
            table_range = f"A{header_row}:{get_column_letter(max_col)}{max_row}"

            table = Table(displayName="ExportTable", ref=table_range)

            style = TableStyleInfo(
                name="TableStyleMedium1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            table.tableStyleInfo = style
            ws.add_table(table)

        # ---- Generous Auto Resize Columns ----
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            generous_padding = 6
            ws.column_dimensions[column_letter].width = max_length + generous_padding

        wb.save(output_path)

        print(f"Exported {self.relation_name} to formatted table {output_path}")

        # ---- Auto Open File ----
        if sys.platform.startswith("darwin"):
            subprocess.call(("open", output_path))
        elif os.name == "nt":
            os.startfile(output_path)
        elif os.name == "posix":
            subprocess.call(("xdg-open", output_path))

